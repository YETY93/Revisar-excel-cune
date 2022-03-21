#!/usr/bin/python
# -*- coding: utf-8 -*-
from subprocess import check_output
from requests import request

import requests
import os
import openpyxl

pathLocal = ""
archivoTemporal = ""
nombreArchivoXLS = ""
extensionXLS = ".xlsx"

listArchivosLocales = []

# toma la ruta actual donde se esta ejecutando
# el scrip
pathLocal = os.getcwd()

# se crea la lista con los archivos que estas dentro
# de la ruta actual
listArchivosLocales = os.listdir(pathLocal)


# Esta funacion recibe le cune de una factura y lo consulta en la DIAN
# returnado 200 si fue aceptada o 500 si no existe el cune ne la DIAN
def consultaEsatado(cune):
    estadoHTTP = ""
    PATH_DIAN = 'https://catalogo-vpfe.dian.gov.co/Document/ShowDocumentToPublic/'

    while (cune != None):
        requestsDIAN = requests.get(PATH_DIAN + cune, timeout=5)

        if (requestsDIAN.status_code == 200):
            estadoHTTP = "REPORTADA_DIAN"
        else:
            estadoHTTP = "PENDIENTE"

        return estadoHTTP


"""
Plantilla a analizar para el estado de las facturas
celda A1 = id_tenant
celda B1 = fac_numero_doc_comercial
Celda C1 = fac_uuid
Celda D1 = fac_estado
Celda E1 = estado

"""


def lecturaArchivXLS(hojaCalculo: str):
    libroExcel = ""
    hojaActiva = ""
    listHojasExcel = []

    INICIO = "id_tenant"
    CELDA_PRINCIPAL = 'A1'

    try:
        libroExcel = openpyxl.load_workbook(hojaCalculo)  # carga de la hoja de excel
        listHojasExcel = libroExcel.sheetnames
        hojaActiva = libroExcel[listHojasExcel[0]]  ## Solo se tomara en cuenta la primera hoja

        columnaEstadoExcel = 'E'
        columnaCuneExcel = 'C'
        ConsecutivoExcel = 2

        if (hojaActiva[CELDA_PRINCIPAL].value == INICIO):
            print("Inicio de análisis: ", hojaCalculo)
            for row in hojaActiva.iter_rows():
                posicionCune = (columnaCuneExcel + str(ConsecutivoExcel))  # se crea la posicion de la celda a leer
                cuneExcel = hojaActiva[posicionCune].value
                estadoTemporalFactura = consultaEsatado(
                    cuneExcel)  # se uriliza la función consultaEsatado para consultar el cune
                posicionEstado = (
                        columnaEstadoExcel + str(ConsecutivoExcel))  # se crea la posicion de la celda a escribir
                hojaActiva[posicionEstado].value = estadoTemporalFactura

                if (estadoTemporalFactura == None):
                    libroExcel.save(hojaCalculo)
                    break

                print("Se ha actualizado la celda ", posicionEstado, " con valor ", estadoTemporalFactura)
                ConsecutivoExcel = (ConsecutivoExcel + 1)
            print("Análisis Finalizado para: ", hojaCalculo)
            # se guardan los cambios
            libroExcel.save(hojaCalculo)
        else:
            print("El archivo no contiene la estructura correcta")

    except Exception as ex:
        print("Archivo corrupto \n posible falla " + str(ex))


# Le asigna los archivos que se encuentran en la ruta "pathLocal"
# a la lista "listArchivosLocales"
def funcionPrinicipal():
    for i in range(0, len(listArchivosLocales)):
        archivoTemporal = listArchivosLocales[i]
        if (extensionXLS in archivoTemporal):
            nombreArchivoXLS = archivoTemporal
            lecturaArchivXLS(nombreArchivoXLS)

    if (nombreArchivoXLS == ""):
        print("No se encontraron archivos compatibles")


# Ejecuta la función principal
funcionPrinicipal()
